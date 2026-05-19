import os
import tkinter as tk
import threading

from tkinter import ttk, filedialog, messagebox
from datetime import datetime

from heidenhain.app_paths import (
    APP_DIR, BILDER_DIR, PROGRAMME_DIR, AUSGABE_DIR,
    ensure_dirs, chdir_to_app_dir
)

from heidenhain.read_pgm import pgm, pgm_with_path
from heidenhain.place_table import tool_in_machin
from heidenhain.tool_table import add_tool_values
from heidenhain.machine_download import machine_download
from heidenhain.standard_tools import remove_standard_tools
from heidenhain.settings import save_json, open_json

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# einmalig Ordner sicherstellen + optional Arbeitsordner setzen
ensure_dirs()
chdir_to_app_dir()


# ---------------------------
# Helfer
# ---------------------------
def string_to_bool(string):
    if string == "Ja":
        return True
    if string == "Nein":
        return False
    return False


def list_files():
    if not PROGRAMME_DIR.exists():
        return []
    return os.listdir(PROGRAMME_DIR)


def test_programs(programs):
    if programs == "Hier Programmnummern eintragen" or programs == "":
        return None
    programs = programs.split(",")
    pgm_list = []
    for pgm_ in programs:
        pgm_ = pgm_.strip()
        if pgm_:
            pgm_list.append(pgm_ + ".h")
    return pgm_list if pgm_list else None


# ---------------------------
# Excel speichern
# ---------------------------
def save_tool_list(tool_list, head_list):
    wb = Workbook()
    ws = wb.active

    darker_blue = "FF00447C"
    blue = "FFDCE6F1"
    dark_blue = "FF0A5C9B"
    white = "FFFFFFFF"

    font_title = Font(name="Arial", size=19, bold=True, color="FFFFFF")
    font_column = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    font_tool_bold = Font(name="Arial", size=16, bold=True, color="000000")
    font_data_bold = Font(name="Arial", size=13, bold=True, color="808080")
    font_cell = Font(name="Arial", size=10)

    border_black = "FF000000"
    thin = Side(border_style="thin", color=border_black)
    thick = Side(border_style="thick", color=border_black)
    thick_all = Border(top=thick, bottom=thick, left=thick, right=thick)

    counter = 0

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 58
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 11
    ws.column_dimensions["E"].width = 5

    last_file = None

    for index, tools_dict in enumerate(tool_list):
        try:
            tool_data_list = list(tools_dict.values())[0]
            file = list(tools_dict.keys())[0]
            last_file = file
            text_lines = list(head_list[index].values())[0]
            head = "\n".join(text_lines)
        except (IndexError, AttributeError) as e:
            print(f"Fehler beim Zugriff auf Daten bei Index {index}: {e}")
            continue

        if index > 0:
            counter += 1

        counter += 1
        ws.merge_cells(f"A{counter}:E{counter}")
        title_cell = ws[f"A{counter}"]
        title_cell.value = f"Werkzeugliste PGM: {file}"
        title_cell.font = font_title
        title_cell.fill = PatternFill("solid", fgColor=darker_blue)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        for col_idx in range(1, 6):
            ws.cell(row=counter, column=col_idx).border = thick_all

        counter += 1
        ws.merge_cells(f"A{counter}:E{counter}")
        head_title = ws[f"A{counter}"]
        head_title.value = "Beschreibung / Aufspannung:"
        head_title.font = font_column
        head_title.fill = PatternFill("solid", fgColor=dark_blue)

        for col_idx in range(1, 6):
            ws.cell(row=counter, column=col_idx).border = thick_all

        counter += 1
        ws.merge_cells(f"A{counter}:E{counter}")
        head_cell = ws[f"A{counter}"]
        head_cell.value = head
        head_cell.font = font_cell
        head_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[counter].height = (font_cell.size+2)*1.11*len(text_lines)

        for col_idx in range(1, 6):
            ws.cell(row=counter, column=col_idx).border = Border(
                left=thick if col_idx == 1 else thin,
                right=thick if col_idx == 5 else thin,
                top=thin,
                bottom=thin
            )

        counter += 1
        headers = ["T-Nr.", "Beschreibung", "L", "R", "K"]

        for col_idx, value in enumerate(headers, 1):
            cell = ws.cell(row=counter, column=col_idx)
            cell.value = value
            cell.font = font_column
            cell.fill = PatternFill("solid", fgColor=dark_blue)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            is_first_col = col_idx == 1
            is_last_col = col_idx == 5

            cell.border = Border(
                top=thick,
                left=thick if is_first_col else thin,
                right=thick if is_last_col else thin,
                bottom=thick
            )

        color_counter = 0

        for tool_detail_dict in tool_data_list:
            counter += 1

            tool_number = list(tool_detail_dict.keys())[0]
            tool_details = list(tool_detail_dict.values())[0]

            color_counter += 1
            row_color = blue if color_counter % 2 == 1 else white

            tool_description = tool_details.get("description", "")

            t_length = tool_details.get("length")
            if t_length is not None:
                if "+" in t_length or "-" in t_length:
                    t_length = t_length[1:]
                t_length = float(t_length)
            else:
                t_length = " "
            tool_length = t_length

            t_radius = tool_details.get("radius")
            if t_radius is not None:
                if "+" in t_radius or "-" in t_radius:
                    t_radius = t_radius[1:]
                t_radius = float(t_radius)
            else:
                t_radius = " "
            tool_radius = t_radius

            cooling_value = tool_details.get("cooling", "")

            ws[f"A{counter}"].value = tool_number
            ws[f"A{counter}"].font = font_tool_bold
            ws[f"A{counter}"].alignment = Alignment(horizontal="center", vertical="center")

            ws[f"B{counter}"].value = tool_description
            ws[f"B{counter}"].font = font_cell
            ws[f"B{counter}"].alignment = Alignment(wrap_text=True, vertical="top")

            ws[f"C{counter}"].value = tool_length if tool_length else " "
            ws[f"C{counter}"].font = font_data_bold if tool_length else Font(name="Arial", size=13)
            ws[f"C{counter}"].alignment = Alignment(horizontal="center", vertical="center")

            ws[f"D{counter}"].value = tool_radius if tool_radius else " "
            ws[f"D{counter}"].font = font_data_bold if tool_radius else Font(name="Arial", size=13)
            ws[f"D{counter}"].alignment = Alignment(horizontal="center", vertical="center")

            ws[f"E{counter}"].value = cooling_value
            ws[f"E{counter}"].font = font_cell
            ws[f"E{counter}"].alignment = Alignment(horizontal="center", vertical="center")

            for col_idx in range(1, 6):
                cell = ws.cell(row=counter, column=col_idx)
                cell.fill = PatternFill("solid", fgColor=row_color)

                is_first_col = col_idx == 1
                is_last_col = col_idx == 5

                cell.border = Border(
                    top=thin,
                    left=thick if is_first_col else thin,
                    right=thick if is_last_col else thin,
                    bottom=thin
                )
        
        now = datetime.now()
        date = now.strftime("%d.%m.%Y")
        time = now.strftime("%H:%M")

        counter += 1
        ws.merge_cells(f"A{counter}:B{counter}")
        date_cell = ws[f"A{counter}"]
        date_cell.value = "  " + date
        date_cell.font = font_column
        date_cell.fill = PatternFill("solid", fgColor=darker_blue)
        date_cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.merge_cells(f"C{counter}:E{counter}")
        time_cell = ws[f"C{counter}"]
        time_cell.value = time + "  "
        time_cell.font = font_column
        time_cell.fill = PatternFill("solid", fgColor=darker_blue)
        time_cell.alignment = Alignment(horizontal="right", vertical="center")

        for col_idx in range(1, 6):
            cell = ws.cell(row=counter, column=col_idx)
            cell.border = Border(
                top=thick,
                left=thick if col_idx == 1 else thin,
                right=thick if col_idx == 5 else thin,
                bottom=thick
            )

        

    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    if not last_file:
        print("Kein Programmname zum Speichern gefunden.")
        return

    base_path = AUSGABE_DIR / f"{last_file}.xlsx"
    path_to_save = base_path

    try:
        wb.save(str(path_to_save))
    except Exception as e:
        print(f"Fehler beim Speichern (Datei eventuell noch geöffnet): {e}")

        i = 1
        while True:
            try:
                alt_name = f"{last_file}{'_' * i}.xlsx"
                path_to_save = AUSGABE_DIR / alt_name
                wb.save(str(path_to_save))
                print(f"Gespeichert als Ausweichdatei: {path_to_save}")
                break
            except Exception as e2:
                i += 1
                if i > 20:
                    return "Fehler beim Speichern. Die Datei konnte nicht erstellt werden.", e2

    try:
        os.startfile(str(path_to_save))
    except Exception as e:
        print(f"Fehler beim Öffnen der Datei: {e}")

    return "Werkzeugliste erfolgreich gespeichert und geöffnet."


# ---------------------------
# Thread Start / Worker
# ---------------------------
def start_tool_thread(frame, controller):
    controller.start_busy(frame.pgm_button, text="⏳ Starte …")
    t = threading.Thread(target=tool_worker, args=(frame, controller), daemon=True)
    t.start()


def start_thread_openpgm(frame, controller):
    controller.start_busy(frame.file_button, text="⏳ Starte …")
    t = threading.Thread(target=open_pgm, args=(frame, controller), daemon=True)
    t.start()


def tool_worker(frame, controller):
    try:
        controller.set_status_safe("⏳ Prüfe Eingaben …")

        machine = frame.selected_machine.get()
        programs = frame.selected_programs.get()

        if machine == "_keine_":
            controller.set_status_safe("⚠️ Fehler: Programmeintrag ungültig.")
            controller.after(0, lambda: controller.error(
                error_title="Fehlerhafter Eintrag",
                error_message="Keine Maschine zum Download von Programmen gewählt.\nMaschine wählen oder Programm manuell öffnen."
            ))
            controller.stop_busy_safe(frame.pgm_button, text="🔵 Bereit.")
            return

        clamp_info = string_to_bool(frame.select_clamp_info.get())
        tool_sort = string_to_bool(frame.select_tool_sort.get())

        pgm_list = test_programs(programs=programs)
        if not pgm_list:
            controller.set_status_safe("⚠️ Fehler: Programmeintrag ungültig.")
            controller.after(0, lambda: controller.error(
                error_title="Fehlerhafter Programmeintrag",
                error_message="Fehlerhaftes oder falsches Programm eingetragen."
            ))
            controller.stop_busy_safe(frame.pgm_button, text="🔵 Bereit.")
            return

        controller.set_status_safe("⏳ Lade Maschine / Programme …")
        subfolder_name = frame.selected_subfolder.get().strip()
        get_error = machine_download(machine=machine, pgm_list=pgm_list, subfolder=subfolder_name)

        if get_error:
            controller.set_status_safe("⚠️ Fehler beim Download/Einlesen.")
            controller.after(0, lambda: controller.error(
                error_title=get_error["error_title"],
                error_message=get_error["error_message"]
            ))
            controller.stop_busy_safe(frame.pgm_button, text="🔵 Bereit.")
            return

        controller.set_status_safe("⏳ Erzeuge Werkzeugliste …")
        _ = list_files()

        tool_list, head_list = pgm(tool_sort=tool_sort, clamp_info=clamp_info, machine=machine)

        controller.set_status_safe("⏳ Prüfe Werkzeuge in Maschine …")
        tool_list = tool_in_machin(tool_list)

        controller.set_status_safe("⏳ Ergänze Werkzeugwerte …")
        tool_list = add_tool_values(tool_list, machine)

        controller.set_status_safe("⏳ Entferne Standardwerkzeuge …")
        tool_list = remove_standard_tools(tool_list=tool_list, machine=machine)

        controller.set_status_safe("⏳ Speichere Excel …")
        save_tool_list(tool_list=tool_list, head_list=head_list)

        controller.stop_busy_safe(frame.pgm_button, text="✅ Fertig: Werkzeugliste erstellt.")

    except Exception as e:
        controller.set_status_safe("⚠️ Unerwarteter Fehler.")
        controller.after(0, lambda: controller.error("Unerwarteter Fehler", str(e)))
        controller.stop_busy_safe(frame.pgm_button, text="🔵 Bereit.")


def open_pgm(frame, controller):
    pgm_path = filedialog.askopenfilename(
        title="Datei auswählen",
        filetypes=[("Heidenhain", "*.h"), ("Alle Dateien", "*.*")]
    )

    if pgm_path == "":
        controller.stop_busy_safe(frame.file_button, text="🔵 Bereit.")
        return

    try:
        controller.set_status_safe("⏳ Prüfe Eingaben …")

        machine = frame.selected_machine.get()

        clamp_info = string_to_bool(frame.select_clamp_info.get())
        tool_sort = string_to_bool(frame.select_tool_sort.get())

        controller.set_status_safe("⏳ Lade Maschine / Programme …")
        get_error = machine_download(machine=machine, pgm_list=None, subfolder="_Bezugsordner_")

        if get_error:
            controller.set_status_safe("⚠️ Fehler beim Download/Einlesen.")
            controller.after(0, lambda: controller.error(
                error_title=get_error["error_title"],
                error_message=get_error["error_message"]
            ))
            controller.stop_busy_safe(frame.file_button, text="🔵 Bereit.")
            return

        controller.set_status_safe("⏳ Erzeuge Werkzeugliste …")

        tool_list, head_list = pgm_with_path(
            tool_sort=tool_sort,
            clamp_info=clamp_info,
            pgm_path=pgm_path,
            machine=machine,
        )

        controller.set_status_safe("⏳ Prüfe Werkzeuge in Maschine …")
        tool_list = tool_in_machin(tool_list)

        controller.set_status_safe("⏳ Ergänze Werkzeugwerte …")
        tool_list = add_tool_values(tool_list, machine)

        controller.set_status_safe("⏳ Entferne Standardwerkzeuge …")
        tool_list = remove_standard_tools(tool_list=tool_list, machine=machine)

        controller.set_status_safe("⏳ Speichere Excel …")
        save_tool_list(tool_list=tool_list, head_list=head_list)

        controller.stop_busy_safe(frame.file_button, text="✅ Fertig: Werkzeugliste erstellt.")

    except Exception as e:
        controller.set_status_safe("⚠️ Unerwarteter Fehler.")
        controller.after(0, lambda: controller.error("Unerwarteter Fehler", str(e)))
        controller.stop_busy_safe(frame.file_button, text="🔵 Bereit.")


# ---------------------------
# GUI
# ---------------------------
class Main(tk.Tk):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        self.geometry("670x720+20+20")
        # self.minsize(670, 710)
        self.title("Tool List Generator")
        # self.resizable(False, False)

        self.load_settings_from_json()

        self.style = ttk.Style(self)
        self.style.theme_use("clam")

        BLUE_DARK = "#0A5C9B"
        BLUE_DARKER = "#00447C"
        BLUE_LIGHT = "#DCE6F1"
        BLUE_SOFT = "#EAF4FF"
        BLUE_BORDER = "#7DB8F0"
        BG = "#F6F8FB"
        TEXT = "#111111"
        MUTED = "#667085"

        ico = BILDER_DIR / "TLG.ico"
        if ico.exists():
            try:
                self.iconbitmap(str(ico))
            except Exception:
                pass

        self.configure(bg=BG, padx=14, pady=12)

        self.style.configure("Card.TFrame", background="white")
        self.style.configure(
            "Header.TLabel",
            background=BLUE_DARKER,
            foreground="white",
            font=("Arial", 13, "bold"),
            padding=(12, 10)
        )
        self.style.configure(
            "App.TLabel",
            background="white",
            foreground=TEXT,
            font=("Arial", 10)
        )
        self.style.configure(
            "Value.TLabel",
            background="white",
            foreground="#1F2937",
            font=("Arial", 10, "bold")
        )
        self.style.configure(
            "Muted.TLabel",
            background="white",
            foreground=MUTED,
            font=("Arial", 9)
        )
        self.style.configure(
            "Section.TLabel",
            background="white",
            foreground=BLUE_DARK,
            font=("Arial", 10, "bold")
        )
        self.style.configure(
            "Blue.TLabelframe",
            background="white",
            bordercolor=BLUE_LIGHT,
            relief="solid"
        )
        self.style.configure(
            "Blue.TLabelframe.Label",
            background="white",
            foreground=BLUE_DARK,
            font=("Arial", 10, "bold")
        )
        self.style.configure(
            "Primary.TButton",
            font=("Arial", 10, "bold"),
            padding=(12, 10),
            background=BLUE_DARK,
            foreground="white",
            borderwidth=0,
            focusthickness=0
        )
        self.style.map(
            "Primary.TButton",
            background=[("active", BLUE_DARKER), ("pressed", BLUE_DARKER), ("disabled", "#9FB7D1")],
            foreground=[("disabled", "#F2F2F2")]
        )
        self.style.configure(
            "Nav.TButton",
            font=("Arial", 10, "bold"),
            padding=(10, 8),
            background=BLUE_SOFT,
            foreground=BLUE_DARKER,
            bordercolor=BLUE_BORDER,
            lightcolor=BLUE_BORDER,
            darkcolor=BLUE_BORDER,
            relief="solid"
        )
        self.style.map(
            "Nav.TButton",
            background=[("active", "#D6EBFF"), ("pressed", "#C6E2FF")],
            foreground=[("active", BLUE_DARKER), ("pressed", BLUE_DARKER)],
            bordercolor=[("focus", "#5AA9F8"), ("active", "#5AA9F8")],
            lightcolor=[("focus", "#A8D6FF"), ("active", "#A8D6FF")],
            darkcolor=[("focus", "#5AA9F8"), ("active", "#5AA9F8")]
        )
        self.style.configure(
            "Danger.TButton",
            font=("Arial", 10, "bold"),
            padding=(12, 10),
            background="#C62828",
            foreground="white",
            borderwidth=0
        )
        self.style.map(
            "Danger.TButton",
            background=[("active", "#A91F1F"), ("pressed", "#8B1A1A")]
        )
        self.style.configure(
            "TEntry",
            fieldbackground="white",
            bordercolor="#B7C9E2",
            lightcolor="#8EC5FF",
            darkcolor="#8EC5FF",
            insertcolor=TEXT,
            padding=6
        )
        self.style.map(
            "TEntry",
            bordercolor=[("focus", "#7DB8F0")],
            lightcolor=[("focus", "#BFE3FF")],
            darkcolor=[("focus", "#7DB8F0")]
        )
        self.style.configure(
            "TCombobox",
            fieldbackground="white",
            bordercolor="#B7C9E2",
            lightcolor="#8EC5FF",
            darkcolor="#8EC5FF",
            arrowsize=14,
            padding=6
        )
        self.style.map(
            "TCombobox",
            bordercolor=[("focus", "#7DB8F0")],
            lightcolor=[("focus", "#BFE3FF")],
            darkcolor=[("focus", "#7DB8F0")]
        )
        self.style.configure(
            "Status.TLabel",
            background=BG,
            foreground="#223",
            font=("Arial", 9)
        )

        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=0)
        self.grid_rowconfigure(2, weight=0)
        self.grid_columnconfigure(0, weight=1)

        self.content_container = ttk.Frame(self)
        self.content_container.grid(row=0, column=0, sticky="nsew", pady=(0, 8))
        self.content_container.grid_rowconfigure(0, weight=1)
        self.content_container.grid_columnconfigure(0, weight=1)

        self.main_frame = MainFrame(self.content_container, self)
        self.main_frame.grid(row=0, column=0, sticky="nsew")

        self.settings_frame = SettingsFrame(self.content_container, self)
        self.settings_frame.grid(row=0, column=0, sticky="nsew")

        self.show_main_frame()

        self.bind("<Return>", lambda _: start_tool_thread(frame=self.main_frame, controller=self))

        ttk.Separator(self).grid(row=1, column=0, sticky="ew", pady=(6, 0))

        self.status_var = tk.StringVar(value="🔵 Bereit.")
        status_container = ttk.Frame(self, padding=(8, 4))
        status_container.grid(row=2, column=0, sticky="ew", pady=(6, 0))
        status_container.grid_columnconfigure(0, weight=1)

        self.status_label = ttk.Label(
            status_container,
            textvariable=self.status_var,
            anchor="w",
            style="Status.TLabel"
        )
        self.status_label.grid(row=0, column=0, sticky="ew")

        self.progress = ttk.Progressbar(status_container, mode="indeterminate", length=160)
        self.progress.grid(row=0, column=1, sticky="e", padx=(10, 0))
        self.progress.grid_remove()

    def set_status_safe(self, text: str):
        self.after(0, lambda: self.status_var.set(text))

    def start_busy(self, button: ttk.Button, text="⏳ Arbeite …"):
        self.status_var.set(text)
        self.progress.grid()
        self.progress.start(12)
        button.state(["disabled"])

    def stop_busy(self, button: ttk.Button, text="🔵 Bereit."):
        self.progress.stop()
        self.progress.grid_remove()
        button.state(["!disabled"])
        self.status_var.set(text)

    def stop_busy_safe(self, button: ttk.Button, text="🔵 Bereit."):
        self.after(0, lambda: self.stop_busy(button, text))

    def get_default_example_data(self):
        return {
            "global_settings": {
                "path": ""
            },
            "machine_settings": {
                "Beispielmaschine": {
                    "ip_address": "192.168.0.10",
                    "reference_folder": r"TNC:\\DMG_160\\",
                    "tool_table_path": r"TNC:\\TOOL.T",
                    "place_table_path": r"TNC:\\TOOL_P.TCH",
                    "subfolders": ["NC", "FRAESEN"],
                    "folder_structure": "Ja",
                    "folder_name_start": "1",
                    "folder_name_end": "1",
                    "default_clamp_info": "Nein",
                    "default_tool_sort": "Ja",
                    "standard_tools": ["99, 999"],
                    "ikz_functions": ["M7, M8, M9"],
                }
            }
        }

    def load_settings_from_json(self):
        data = open_json()

        if data is None:
            data = self.get_default_example_data()

        self.global_settings = data.get("global_settings", {"path": ""})
        self.machine_settings = data.get("machine_settings", {})

        if not self.machine_settings:
            self.machine_settings = self.get_default_example_data()["machine_settings"]

    def save_all_settings_to_json(self):
        data = {
            "global_settings": self.global_settings,
            "machine_settings": self.machine_settings
        }
        save_json(data)

    def show_main_frame(self):
        self.main_frame.refresh_machine_dropdown()
        self.main_frame.refresh_subfolder_dropdown()
        self.main_frame.apply_machine_defaults()
        self.main_frame.load_machine_details()
        self.main_frame.tkraise()

    def show_settings_frame(self):
        self.settings_frame.refresh_dropdown()
        self.settings_frame.tkraise()

    def get_global_path(self):
        return self.global_settings.get("path", "")

    def get_available_machine_names(self):
        machine_names = sorted(self.machine_settings.keys())

        if "_keine_" in machine_names:
            machine_names.remove("_keine_")

        desired_order = ["DMU_80P", "DMG_125", "Kekeisen", "POS"]
        ordered = [name for name in desired_order if name in machine_names]

        if "_keine_" not in machine_names:
            machine_names.append("_keine_")

        for name in machine_names:
            if name not in ordered:
                ordered.append(name)

        return ordered

    def get_machine_setting(self, machine_name: str):
        return self.machine_settings.get(machine_name, {})

    def get_subfolders_for_machine(self, machine_name: str):
        if not machine_name or machine_name == "_keine_":
            return ["_Bezugsordner_"]

        machine_data = self.get_machine_setting(machine_name)
        stored_subfolders = machine_data.get("subfolders", [])

        values = ["_Bezugsordner_"]
        for item in stored_subfolders:
            item = item.strip()
            if item and item not in values:
                values.append(item)

        return values

    def get_machine_defaults(self, machine_name: str):
        if not machine_name or machine_name == "_keine_":
            return {
                "clamp_info": "Nein",
                "tool_sort": "Nein"
            }

        machine_data = self.get_machine_setting(machine_name)
        return {
            "clamp_info": machine_data.get("default_clamp_info", "Nein"),
            "tool_sort": machine_data.get("default_tool_sort", "Nein")
        }

    def get_machine_details(self, machine_name: str):
        if not machine_name or machine_name == "_keine_":
            return {
                "ip": "",
                "reference_folder": "",
                "tool_table_path": "",
                "place_table_path": "",
                "folder_structure": "Nein",
                "folder_name_start": "1",
                "folder_name_end": "1"
            }

        machine_data = self.get_machine_setting(machine_name)

        return {
            "ip": machine_data.get("ip_address", ""),
            "reference_folder": machine_data.get("reference_folder", ""),
            "tool_table_path": machine_data.get("tool_table_path", ""),
            "place_table_path": machine_data.get("place_table_path", ""),
            "folder_structure": machine_data.get("folder_structure", "Nein"),
            "folder_name_start": machine_data.get("folder_name_start", "1"),
            "folder_name_end": machine_data.get("folder_name_end", "1")
        }

    def build_dynamic_order_path(self, machine_name: str, subfolder_name: str, program_text: str):
        details = self.get_machine_details(machine_name)
        base_folder = details["reference_folder"].strip()
        folder_structure = details["folder_structure"]
        start_value = details["folder_name_start"]
        end_value = details["folder_name_end"]
        if not base_folder:
            return ""

        final_path = base_folder

        if folder_structure == "Ja":
            first_program = ""
            if program_text.strip():
                first_program = program_text.split(",")[0].strip()

            if first_program:
                try:
                    start_index = int(start_value) - 1
                    end_index = int(end_value)

                    if start_index < 0:
                        start_index = 0

                    derived_folder = first_program[start_index:end_index].strip()
                    if not first_program.startswith("Hier Programmnummern eintragen"):
                        if derived_folder:
                            final_path = os.path.join(final_path, derived_folder)
                except ValueError:
                    pass

        if subfolder_name and subfolder_name != "_Bezugsordner_":
            final_path = os.path.join(base_folder, subfolder_name)

        return final_path.replace("\\", "/")

    def save_global_path(self, path_value: str):
        self.global_settings["path"] = path_value.strip()
        self.save_all_settings_to_json()

    def save_machine_settings(
        self,
        machine_name: str,
        ip_value: str,
        reference_folder_value: str,
        tool_table_path: str,
        place_table_path: str,
        subfolder_list: list,
        folder_structure_value: str,
        folder_name_start: str,
        folder_name_end: str,
        standard_tool_list: list,
        ikz_function_list: list,
        clamp_info_value: str,
        tool_sort_value: str
    ):
        if machine_name == "_keine_":
            return

        if machine_name not in self.machine_settings:
            self.machine_settings[machine_name] = {}

        self.machine_settings[machine_name]["ip_address"] = ip_value.strip()
        self.machine_settings[machine_name]["reference_folder"] = reference_folder_value.strip()
        self.machine_settings[machine_name]["tool_table_path"] = tool_table_path.strip()
        self.machine_settings[machine_name]["place_table_path"] = place_table_path.strip()
        self.machine_settings[machine_name]["subfolders"] = subfolder_list
        self.machine_settings[machine_name]["folder_structure"] = folder_structure_value.strip()
        self.machine_settings[machine_name]["folder_name_start"] = folder_name_start.strip()
        self.machine_settings[machine_name]["folder_name_end"] = folder_name_end.strip()
        self.machine_settings[machine_name]["standard_tools"] = standard_tool_list
        self.machine_settings[machine_name]["ikz_functions"] = ikz_function_list
        self.machine_settings[machine_name]["default_clamp_info"] = clamp_info_value.strip()
        self.machine_settings[machine_name]["default_tool_sort"] = tool_sort_value.strip()
        self.save_all_settings_to_json()

    def delete_machine_settings(self, machine_name: str):
        if not machine_name or machine_name == "_keine_":
            return

        if machine_name in self.machine_settings:
            del self.machine_settings[machine_name]
            self.save_all_settings_to_json()

    def error(self, error_title, error_message):
        top = tk.Toplevel(self)
        top.title(error_title)
        top.geometry("420x150+80+100")
        top.resizable(False, False)
        top.configure(bg="#F6F8FB")

        content = ttk.Frame(top, padding=12)
        content.pack(fill="both", expand=True)

        tk.Label(content, image="::tk::icons::error").pack(side="top", pady=(0, 6))
        ttk.Label(content, text=error_message, wraplength=380, justify="center").pack(side="top", pady=(0, 10))
        ttk.Button(content, text="OK", width=12, command=top.destroy, style="Primary.TButton").pack(side="top")


class MainFrame(ttk.Frame):
    def __init__(self, container, controller, **kwargs):
        super().__init__(container, **kwargs)
        self.controller = controller

        self.selected_machine = tk.StringVar(value="_keine_")
        self.selected_subfolder = tk.StringVar(value="_Bezugsordner_")
        self.selected_programs = tk.StringVar(value="Hier Programmnummern eintragen")
        self.select_clamp_info = tk.StringVar(value="Nein")
        self.select_tool_sort = tk.StringVar(value="Nein")

        self.display_ip_var = tk.StringVar()
        self.display_reference_folder_var = tk.StringVar()
        self.display_tool_table_var = tk.StringVar()
        self.display_place_table_var = tk.StringVar()
        self.display_folder_structure_var = tk.StringVar(value="Nein")
        self.display_global_path_var = tk.StringVar()

        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        card = ttk.Frame(self, style="Card.TFrame", padding=18)
        card.grid(row=0, column=0, sticky="nsew")
        card.grid_columnconfigure(0, weight=1)
        card.grid_columnconfigure(1, weight=1)

        header_frame = ttk.Frame(card)
        header_frame.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 14))
        header_frame.grid_columnconfigure(0, weight=1)

        ttk.Label(header_frame, text="Werkzeuglisten Generator", style="Header.TLabel").grid(row=0, column=0, sticky="ew")

        ttk.Button(
            header_frame,
            text="⚙ Einstellungen",
            style="Nav.TButton",
            command=self.controller.show_settings_frame
        ).grid(row=0, column=1, padx=(10, 0))

        ttk.Label(card, text="Maschine:", style="Section.TLabel").grid(row=1, column=0, sticky="w", padx=(0, 8))
        ttk.Label(card, text="Unterordner:", style="Section.TLabel").grid(row=1, column=1, sticky="w", padx=(8, 0))

        self.machine_combobox = ttk.Combobox(card, textvariable=self.selected_machine, state="readonly")
        self.machine_combobox.grid(row=2, column=0, sticky="ew", pady=(4, 12), padx=(0, 8))
        self.machine_combobox.bind("<<ComboboxSelected>>", self.on_machine_selected)

        self.subfolder_combobox = ttk.Combobox(card, textvariable=self.selected_subfolder, state="readonly")
        self.subfolder_combobox.grid(row=2, column=1, sticky="ew", pady=(4, 12), padx=(8, 0))
        self.subfolder_combobox.bind("<<ComboboxSelected>>", self.on_subfolder_selected)

        self.refresh_machine_dropdown()

        available_machines = self.controller.get_available_machine_names()
        if available_machines:
            self.selected_machine.set(available_machines[0])
        else:
            self.selected_machine.set("_keine_")

        self.refresh_subfolder_dropdown()
        self.selected_subfolder.set("_Bezugsordner_")
        self.apply_machine_defaults()
        self.load_machine_details()

        options = ttk.Labelframe(card, text="Optionen", style="Blue.TLabelframe", padding=12)
        options.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(0, 12))
        options.grid_columnconfigure(0, weight=1)
        options.grid_columnconfigure(1, weight=1)

        ttk.Label(options, text="Mit Spann-Notizen:", style="App.TLabel").grid(row=0, column=0, sticky="w", pady=3)
        ttk.Combobox(
            options,
            textvariable=self.select_clamp_info,
            state="readonly",
            values=("Ja", "Nein")
        ).grid(row=1, column=0, sticky="ew", padx=(2, 10), pady=2)

        ttk.Label(options, text="Werkzeug sortieren:", style="App.TLabel").grid(row=0, column=1, sticky="w", pady=3)
        ttk.Combobox(
            options,
            textvariable=self.select_tool_sort,
            state="readonly",
            values=("Ja", "Nein")
        ).grid(row=1, column=1, sticky="ew", padx=(10, 2), pady=2)

        ttk.Label(card, text="Programme (kommagetrennt):", style="Section.TLabel").grid(row=4, column=0, columnspan=2, sticky="w")

        pgm_entry = ttk.Entry(card, textvariable=self.selected_programs, font=("Arial", 10))
        pgm_entry.grid(row=5, column=0, columnspan=2, sticky="ew", pady=(4, 14))
        pgm_entry.bind("<Button-1>", self.clear_pgm_entry)
        pgm_entry.bind("<KeyRelease>", self.on_program_entry_changed)

        button_frame = ttk.Frame(card, style="Card.TFrame")
        button_frame.grid(row=6, column=0, columnspan=2, sticky="ew")
        button_frame.grid_columnconfigure(0, weight=1)
        button_frame.grid_columnconfigure(1, weight=1)

        self.pgm_button = ttk.Button(
            button_frame,
            text="WKZ-Liste generieren",
            style="Primary.TButton",
            command=lambda: start_tool_thread(frame=self, controller=controller)
        )
        self.pgm_button.grid(row=0, column=0, sticky="ew", padx=(0, 8))

        self.file_button = ttk.Button(
            button_frame,
            text="Programm öffnen",
            style="Primary.TButton",
            command=lambda: start_thread_openpgm(frame=self, controller=controller)
        )
        self.file_button.grid(row=0, column=1, sticky="ew", padx=(8, 0))

        info_box = ttk.Labelframe(card, text="Aktive Einstellungen", style="Blue.TLabelframe", padding=12)
        info_box.grid(row=7, column=0, columnspan=2, sticky="ew", pady=(14, 0))
        info_box.grid_columnconfigure(1, weight=1)

        ttk.Label(info_box, text="Maschine:", style="App.TLabel").grid(row=0, column=0, sticky="nw", pady=3)
        ttk.Label(info_box, textvariable=self.selected_machine, style="Value.TLabel").grid(row=0, column=1, sticky="nw", pady=3)

        ttk.Label(info_box, text="Globaler Pfad:", style="App.TLabel").grid(row=2, column=0, sticky="nw", pady=3)
        ttk.Label(
            info_box,
            textvariable=self.display_global_path_var,
            style="Value.TLabel",
            wraplength=560,
            justify="left"
        ).grid(row=2, column=1, sticky="nw", pady=3)

        ttk.Label(info_box, text="IP-Adresse:", style="App.TLabel").grid(row=3, column=0, sticky="nw", pady=3)
        ttk.Label(info_box, textvariable=self.display_ip_var, style="Value.TLabel").grid(row=3, column=1, sticky="nw", pady=3)

        ttk.Label(info_box, text="Aktiver-Pfad:", style="App.TLabel").grid(row=4, column=0, sticky="nw", pady=3)
        ttk.Label(
            info_box,
            textvariable=self.display_reference_folder_var,
            style="Value.TLabel",
            wraplength=560,
            justify="left"
        ).grid(row=4, column=1, sticky="nw", pady=3)

        ttk.Label(info_box, text="Pfad zur Werkzeugtabelle:", style="App.TLabel").grid(row=5, column=0, sticky="nw", pady=3)
        ttk.Label(
            info_box,
            textvariable=self.display_tool_table_var,
            style="Value.TLabel",
            wraplength=560,
            justify="left"
        ).grid(row=5, column=1, sticky="nw", pady=3)

        ttk.Label(info_box, text="Pfad zur Platztabelle:", style="App.TLabel").grid(row=6, column=0, sticky="nw", pady=3)
        ttk.Label(
            info_box,
            textvariable=self.display_place_table_var,
            style="Value.TLabel",
            wraplength=560,
            justify="left"
        ).grid(row=6, column=1, sticky="nw", pady=3)

        ttk.Label(info_box, text="Ordner-Struktur:", style="App.TLabel").grid(row=7, column=0, sticky="nw", pady=3)
        ttk.Label(info_box, textvariable=self.display_folder_structure_var, style="Value.TLabel").grid(row=7, column=1, sticky="nw", pady=3)

        ttk.Label(
            info_box,
            text="Beispiel: Programm '12345' mit Zeichen von 1 bis 1 ergibt den Ordnernamen '1'.",
            style="Muted.TLabel",
            wraplength=700,
            justify="left"
        ).grid(row=8, column=0, columnspan=2, sticky="w", pady=(10, 0))

    def refresh_machine_dropdown(self):
        values = self.controller.get_available_machine_names()
        self.machine_combobox["values"] = values
        if self.selected_machine.get() not in values:
            self.selected_machine.set("_keine_")

    def refresh_subfolder_dropdown(self):
        selected_machine = self.selected_machine.get().strip()
        subfolders = self.controller.get_subfolders_for_machine(selected_machine)
        if not subfolders:
            subfolders = ["_Bezugsordner_"]

        self.subfolder_combobox["values"] = subfolders

        current = self.selected_subfolder.get().strip()
        if current not in subfolders:
            self.selected_subfolder.set("_Bezugsordner_")

    def apply_machine_defaults(self):
        defaults = self.controller.get_machine_defaults(self.selected_machine.get().strip())
        self.select_clamp_info.set(defaults["clamp_info"])
        self.select_tool_sort.set(defaults["tool_sort"])

    def load_machine_details(self):
        details = self.controller.get_machine_details(self.selected_machine.get().strip())
        self.display_global_path_var.set(self.controller.get_global_path())
        self.display_ip_var.set(details["ip"])
        self.display_tool_table_var.set(details["tool_table_path"])
        self.display_place_table_var.set(details["place_table_path"])
        self.display_folder_structure_var.set(details["folder_structure"])
        self.update_dynamic_reference_folder()

    def update_dynamic_reference_folder(self):
        dynamic_path = self.controller.build_dynamic_order_path(
            machine_name=self.selected_machine.get().strip(),
            subfolder_name=self.selected_subfolder.get().strip(),
            program_text=self.selected_programs.get().strip()
        )
        self.display_reference_folder_var.set(dynamic_path)

    def on_machine_selected(self, event=None):
        self.refresh_subfolder_dropdown()
        self.apply_machine_defaults()
        self.load_machine_details()

    def on_subfolder_selected(self, event=None):
        self.update_dynamic_reference_folder()

    def on_program_entry_changed(self, event=None):
        self.update_dynamic_reference_folder()

    def clear_pgm_entry(self, event=None):
        if self.selected_programs.get() == "Hier Programmnummern eintragen":
            self.selected_programs.set("")
            self.update_dynamic_reference_folder()


class SettingsFrame(ttk.Frame):
    def __init__(self, container, controller, **kwargs):
        super().__init__(container, **kwargs)
        self.controller = controller

        # Variablen
        self.selected_setting = tk.StringVar()
        self.tnccmd_path_var = tk.StringVar()
        self.name_var = tk.StringVar()
        self.ip_var = tk.StringVar()
        self.reference_folder_var = tk.StringVar()
        self.tool_table_path_var = tk.StringVar()
        self.place_table_path_var = tk.StringVar()
        self.unterordner_var = tk.StringVar()
        self.ordner_struktur_var = tk.StringVar(value="Nein")
        self.folder_name_start_var = tk.StringVar(value="1")
        self.folder_name_end_var = tk.StringVar(value="1")
        self.standard_tool_var = tk.StringVar()
        self.ikz_var = tk.StringVar()
        self.default_clamp_info_var = tk.StringVar(value="Nein")
        self.default_tool_sort_var = tk.StringVar(value="Nein")

        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # === Haupt-Card ===
        card = ttk.Frame(self, style="Card.TFrame", padding=18)
        card.grid(row=0, column=0, sticky="nsew")
        card.grid_columnconfigure(1, weight=1)
        card.grid_columnconfigure(2, weight=0)
        card.grid_rowconfigure(6, weight=1)

        # Header
        header_frame = ttk.Frame(card)
        header_frame.grid(row=0, column=0, columnspan=3, sticky="ew", pady=(0, 14))
        header_frame.grid_columnconfigure(0, weight=1)

        ttk.Label(header_frame, text="Einstellungen", style="Header.TLabel").grid(row=0, column=0, sticky="ew")

        ttk.Button(
            header_frame,
            text="← Zurück",
            style="Nav.TButton",
            command=self.controller.show_main_frame
        ).grid(row=0, column=1, padx=(10, 0))

        # TNCcmd
        ttk.Label(card, text="Pfad zur TNCcmd.exe:", style="Section.TLabel").grid(
            row=1, column=0, columnspan=3, sticky="w"
        )

        ttk.Entry(card, textvariable=self.tnccmd_path_var).grid(
            row=2, column=0, columnspan=2, sticky="ew", pady=(4, 10), padx=(0, 8)
        )

        ttk.Button(
            card,
            text="Datei wählen",
            style="Nav.TButton",
            command=self.select_file
        ).grid(row=2, column=2, sticky="ew", pady=(4, 10))

        # Gespeicherte Maschine
        ttk.Label(card, text="Gespeicherte Maschine:", style="Section.TLabel").grid(
            row=3, column=0, columnspan=3, sticky="w"
        )

        self.settings_combobox = ttk.Combobox(card, textvariable=self.selected_setting, state="readonly")
        self.settings_combobox.grid(row=4, column=0, columnspan=3, sticky="ew", pady=(4, 12))
        self.settings_combobox.bind("<<ComboboxSelected>>", self.load_selected_setting)

        # === Umschalt-Buttons ===
        switch_row = ttk.Frame(card, style="Card.TFrame")
        switch_row.grid(row=5, column=0, columnspan=3, sticky="ew", pady=(0, 10))
        switch_row.grid_columnconfigure(0, weight=1)
        switch_row.grid_columnconfigure(1, weight=1)

        # Container für Seiten
        pages_frame = ttk.Frame(card, style="Card.TFrame")
        pages_frame.grid(row=6, column=0, columnspan=3, sticky="nsew")
        pages_frame.grid_columnconfigure(0, weight=1)
        pages_frame.grid_rowconfigure(0, weight=1)

        # Zwei Frames (Seiten)
        machine_frame = ttk.Frame(pages_frame, style="Card.TFrame")
        list_frame = ttk.Frame(pages_frame, style="Card.TFrame")

        for frame in (machine_frame, list_frame):
            frame.grid(row=0, column=0, sticky="nsew")
            frame.grid_columnconfigure(1, weight=1)
            frame.grid_columnconfigure(2, weight=0)

        # Buttons (nachdem Frames existieren!)
        ttk.Button(
            switch_row,
            text="Maschinen Einstellungen",
            style="Nav.TButton",
            command=lambda: self.show_settings_page(machine_frame)
        ).grid(row=0, column=0, sticky="ew", padx=(0, 6), pady=8)

        ttk.Button(
            switch_row,
            text="WKZ-Listen Einstellungen",
            style="Nav.TButton",
            command=lambda: self.show_settings_page(list_frame)
        ).grid(row=0, column=1, sticky="ew", padx=(6, 0), pady=8)

        # =========================================================
        # FRAME 1 – Maschine
        # =========================================================
        ttk.Label(machine_frame, text="Name:", style="App.TLabel").grid(row=0, column=0, sticky="w", pady=3)
        ttk.Entry(machine_frame, textvariable=self.name_var).grid(row=0, column=1, columnspan=2, sticky="ew", pady=3)

        ttk.Label(machine_frame, text="IP-Adresse:", style="App.TLabel").grid(row=1, column=0, sticky="w", pady=3)
        ttk.Entry(machine_frame, textvariable=self.ip_var).grid(row=1, column=1, columnspan=2, sticky="ew", pady=3)

        ttk.Label(machine_frame, text="Bezugsordner-Pfad:", style="App.TLabel").grid(row=2, column=0, sticky="w", pady=3)
        ttk.Entry(machine_frame, textvariable=self.reference_folder_var).grid(row=2, column=1, columnspan=2, sticky="ew", pady=3)

        ttk.Label(machine_frame, text="Pfad zur Werkzeugtabelle:", style="App.TLabel").grid(row=3, column=0, sticky="w", pady=3)
        ttk.Entry(machine_frame, textvariable=self.tool_table_path_var).grid(row=3, column=1, columnspan=2, sticky="ew", pady=3)

        ttk.Label(machine_frame, text="Pfad zur Platztabelle:", style="App.TLabel").grid(row=4, column=0, sticky="w", pady=3)
        ttk.Entry(machine_frame, textvariable=self.place_table_path_var).grid(row=4, column=1, columnspan=2, sticky="ew", pady=3)

        ttk.Label(machine_frame, text="Unterordner (kommagetrennt):", style="App.TLabel").grid(row=5, column=0, sticky="w", pady=3)
        ttk.Entry(machine_frame, textvariable=self.unterordner_var).grid(row=5, column=1, columnspan=2, sticky="ew", pady=3)

        ttk.Label(machine_frame, text="Ordner-Struktur aktiv:", style="App.TLabel").grid(row=6, column=0, sticky="w", pady=3)
        ttk.Combobox(
            machine_frame,
            textvariable=self.ordner_struktur_var,
            state="readonly",
            values=("Ja", "Nein")
        ).grid(row=6, column=1, columnspan=2, sticky="ew", pady=3)

        ttk.Label(machine_frame, text="Ordner Zeichen:", style="App.TLabel").grid(row=7, column=0, sticky="w", pady=3)

        zeichen_frame = ttk.Frame(machine_frame, style="Card.TFrame")
        zeichen_frame.grid(row=7, column=1, columnspan=2, sticky="ew", pady=3)

        ttk.Label(zeichen_frame, text="von", style="App.TLabel").grid(row=0, column=0, padx=(0, 6), sticky="w")
        ttk.Combobox(
            zeichen_frame,
            textvariable=self.folder_name_start_var,
            state="readonly",
            width=6,
            values=[str(i) for i in range(1, 21)]
        ).grid(row=0, column=1, padx=(0, 12), sticky="w")

        ttk.Label(zeichen_frame, text="bis", style="App.TLabel").grid(row=0, column=2, padx=(0, 6), sticky="w")
        ttk.Combobox(
            zeichen_frame,
            textvariable=self.folder_name_end_var,
            state="readonly",
            width=6,
            values=[str(i) for i in range(1, 21)]
        ).grid(row=0, column=3, sticky="w")

        button_row = ttk.Frame(card, style="Card.TFrame")
        button_row.grid(row=7, column=0, columnspan=3, sticky="ew", pady=(14, 0))
        button_row.grid_columnconfigure(0, weight=1)
        button_row.grid_columnconfigure(1, weight=1)

        ttk.Button(button_row, text="Speichern", style="Primary.TButton", command=self.save_settings).grid(
            row=0, column=0, sticky="ew", padx=(0, 8)
        )

        ttk.Button(button_row, text="🗑 Löschen", style="Danger.TButton", command=self.delete_selected_machine).grid(
            row=0, column=1, sticky="ew", padx=(8, 0)
        )

        # =========================================================
        # FRAME 2 – Weitere Einstellungen
        # =========================================================

        ttk.Label(list_frame, text="Standard-Werkzeuge:", style="App.TLabel").grid(row=0, column=0, sticky="w", pady=3)
        ttk.Entry(list_frame, textvariable=self.standard_tool_var).grid(row=0, column=1, columnspan=2, sticky="ew", pady=3)
        
        ttk.Label(list_frame, text="IKZ-Funktionen:", style="App.TLabel").grid(row=1, column=0, sticky="w", pady=3)
        ttk.Entry(list_frame, textvariable=self.ikz_var).grid(row=1, column=1, columnspan=2, sticky="ew", pady=3)

        ttk.Label(list_frame, text="Default Mit Spann-Notizen:", style="App.TLabel").grid(row=3, column=0, sticky="w", pady=3)
        ttk.Combobox(
            list_frame,
            textvariable=self.default_clamp_info_var,
            state="readonly",
            values=("Ja", "Nein")
        ).grid(row=3, column=1, columnspan=2, sticky="ew", pady=3)

        ttk.Label(list_frame, text="Default Werkzeug sortieren:", style="App.TLabel").grid(row=4, column=0, sticky="w", pady=3)
        ttk.Combobox(
            list_frame,
            textvariable=self.default_tool_sort_var,
            state="readonly",
            values=("Ja", "Nein")
        ).grid(row=4, column=1, columnspan=2, sticky="ew", pady=3)

        
        self.refresh_dropdown()

        # Startansicht
        machine_frame.tkraise()

    def show_settings_page(self, frame):
        frame.tkraise()

    def select_file(self):
        dateipfad = filedialog.askopenfilename(
            title="Datei auswählen",
            filetypes=[("Alle Dateien", "*.*")]
        )
        if dateipfad:
            self.tnccmd_path_var.set(dateipfad)

    def refresh_dropdown(self):
        values = [name for name in self.controller.get_available_machine_names() if name != "_keine_"]
        if not values:
            values = [""]

        self.settings_combobox["values"] = values
        self.tnccmd_path_var.set(self.controller.get_global_path())

        current = self.selected_setting.get().strip()
        if current in values:
            return

        self.selected_setting.set(values[0] if values[0] else "")
        if self.selected_setting.get():
            self.load_selected_setting()
        else:
            self.clear_fields()

    def clear_fields(self):
        self.name_var.set("")
        self.ip_var.set("")
        self.reference_folder_var.set("")
        self.tool_table_path_var.set("")
        self.place_table_path_var.set("")
        self.unterordner_var.set("")
        self.ordner_struktur_var.set("Nein")
        self.folder_name_start_var.set("1")
        self.folder_name_end_var.set("1")
        self.standard_tool_var.set("")
        self.ikz_var.set("")
        self.default_clamp_info_var.set("Nein")
        self.default_tool_sort_var.set("Nein")

    def load_selected_setting(self, *args):
        machine_name = self.selected_setting.get().strip()
        if not machine_name or machine_name == "_keine_":
            self.clear_fields()
            return

        data = self.controller.get_machine_setting(machine_name)

        self.name_var.set(machine_name)
        self.tnccmd_path_var.set(self.controller.get_global_path())
        self.ip_var.set(data.get("ip_address", ""))
        self.reference_folder_var.set(data.get("reference_folder", ""))
        self.tool_table_path_var.set(data.get("tool_table_path", ""))
        self.place_table_path_var.set(data.get("place_table_path", ""))
        self.unterordner_var.set(", ".join(data.get("subfolders", [])))
        self.ordner_struktur_var.set(data.get("folder_structure", "Nein"))
        self.folder_name_start_var.set(data.get("folder_name_start", "1"))
        self.folder_name_end_var.set(data.get("folder_name_end", "1"))
        self.standard_tool_var.set(", ".join(data.get("standard_tools", [])))
        self.ikz_var.set(", ".join(data.get("ikz_functions", [])))
        self.default_clamp_info_var.set(data.get("default_clamp_info", "Nein"))
        self.default_tool_sort_var.set(data.get("default_tool_sort", "Nein"))

    def save_settings(self):
        machine_name = self.name_var.get().strip()

        if not machine_name:
            self.controller.error("Fehler", "Bitte einen Namen eingeben.")
            return

        if machine_name == "_keine_":
            self.controller.error("Fehler", "Für '_keine_' dürfen keine Einstellungen gespeichert werden.")
            return

        start_value = int(self.folder_name_start_var.get())
        end_value = int(self.folder_name_end_var.get())

        if start_value > end_value:
            self.controller.error("Fehler", "„Ordnername aus Zeichen von“ darf nicht größer als „bis“ sein.")
            return

        subfolder_list = [
            item.strip()
            for item in self.unterordner_var.get().split(",")
            if item.strip() and item.strip().lower() != "_bezugsordner_"
        ]

        standard_tool_list = [
            item.strip()
            for item in self.standard_tool_var.get().split(",")
        ]

        ikz_function_list = [
            item.strip()
            for item in self.ikz_var.get().split(",")
        ]


        self.controller.save_global_path(self.tnccmd_path_var.get())

        self.controller.save_machine_settings(
            machine_name=machine_name,
            ip_value=self.ip_var.get(),
            reference_folder_value=self.reference_folder_var.get(),
            tool_table_path=self.tool_table_path_var.get(),
            place_table_path=self.place_table_path_var.get(),
            subfolder_list=subfolder_list,
            folder_structure_value=self.ordner_struktur_var.get(),
            folder_name_start=self.folder_name_start_var.get(),
            folder_name_end=self.folder_name_end_var.get(),
            standard_tool_list=standard_tool_list,
            ikz_function_list=ikz_function_list,
            clamp_info_value=self.default_clamp_info_var.get(),
            tool_sort_value=self.default_tool_sort_var.get()
        )

        self.refresh_dropdown()
        self.selected_setting.set(machine_name)

        self.controller.main_frame.refresh_machine_dropdown()
        self.controller.main_frame.selected_machine.set(machine_name)
        self.controller.main_frame.refresh_subfolder_dropdown()
        self.controller.main_frame.selected_subfolder.set("_Bezugsordner_")
        self.controller.main_frame.apply_machine_defaults()
        self.controller.main_frame.load_machine_details()

        self.controller.status_var.set(f"🟢 Einstellung '{machine_name}' gespeichert.")

    def delete_selected_machine(self):
        machine_name = self.selected_setting.get().strip()

        if not machine_name:
            self.controller.error("Fehler", "Es ist keine Maschine ausgewählt.")
            return

        if machine_name == "_keine_":
            self.controller.error("Fehler", "'_keine_' kann nicht gelöscht werden.")
            return

        answer = messagebox.askyesno(
            "Maschine löschen",
            f"Möchtest du die Einstellung '{machine_name}' wirklich löschen?"
        )

        if not answer:
            return

        self.controller.delete_machine_settings(machine_name)

        values = [name for name in self.controller.get_available_machine_names() if name != "_keine_"]
        self.settings_combobox["values"] = values

        if values:
            self.selected_setting.set(values[0])
            self.load_selected_setting()
        else:
            self.selected_setting.set("")
            self.clear_fields()

        self.controller.main_frame.refresh_machine_dropdown()
        self.controller.main_frame.selected_machine.set("_keine_")
        self.controller.main_frame.refresh_subfolder_dropdown()
        self.controller.main_frame.selected_subfolder.set("_Bezugsordner_")
        self.controller.main_frame.apply_machine_defaults()
        self.controller.main_frame.load_machine_details()

        self.controller.status_var.set(f"🗑 Einstellung '{machine_name}' wurde gelöscht.")


if __name__ == "__main__":
    app = Main()
    app.mainloop()